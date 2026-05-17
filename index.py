import pandas as pd
import numpy as np
import random
from datetime import datetime, timedelta

# Seed for reproducibility
np.random.seed(42)
random.seed(42)

# Generate realistic lookup data for Myanmar E-commerce
products = {
    'Electronics': [
        ('Xiaomi Redmi Note 13', 580000),
        ('Samsung Galaxy A25', 720000),
        ('Anker PowerBank 20000mAh', 85000),
        ('Sony WH-CH520 Wireless Headphones', 165000),
        ('Apple iPhone 15 Pro', 4200000)
    ],
    'Fashion & Apparel': [
        ('Myanmar Traditional Longyi (Silk)', 45000),
        ('Unisex Oversized Graphic T-Shirt', 18000),
        ('Crocs Classic Clogs', 120000),
        ('Linen Casual Shirt', 28000),
        ('Ladies Summer Dress', 35000)
    ],
    'Cosmetics & Beauty': [
        ('Shiseido Anessa Sunscreen', 75000),
        ('COSRX Snail Mucin Essence', 48000),
        ('Natural Tanaka Face Powder', 8500),
        ('Innisfree Green Tea Seed Serum', 62000),
        ('Maybelline Fit Me Foundation', 29000)
    ],
    'Home & Groceries': [
        ('Premier Coffee 3-in-1 (30 Pack)', 7500),
        ('Royal Myanmar Tea Mix', 8200),
        ('Electric Hot Pot Multi-Cooker', 65000),
        ('Stainless Steel Water Bottle 1L', 16000),
        ('Smart LED Bulb 9W', 12000)
    ]
}

townships_by_state = {
    'Yangon': ['Hlaing', 'Kamayut', 'Latha', 'Tamwe', 'Mayangone', 'Insein', 'North Dagon', 'South Okkalapa'],
    'Mandalay': ['Chanayethazan', 'Maha Aungmye', 'Pyigyidagun', 'Chanmyathazi', 'Aungmyethazan'],
    'Shan State': ['Taunggyi', 'Kalaw', 'Lashio', 'Muse'],
    'Mon State': ['Mawlamyine', 'Mudon', 'Thaton'],
    'Bago': ['Bago', 'Pyay', 'Taungoo'],
    'Naypyitaw': ['Zabuthiri', 'Ottarathiri', 'Dekkhinathiri']
}

payment_methods = ['KBZPay', 'WaveMoney', 'CB Pay', 'AYAPay', 'Cash on Delivery (COD)']
channels = ['Facebook Messenger', 'Viber Shop', 'Website', 'TikTok Shop', 'ShweShop Mobile App']

# Generate clean base data first, then inject dirtiness
num_rows = 150
start_date = datetime(2026, 1, 1)

data = []

for i in range(1, num_rows + 1):
    order_id = f"ORD-2026-{1000 + i}"

    # Select Category & Product
    category = random.choice(list(products.keys()))
    product_name, base_price = random.choice(products[category])

    # Quantity & Calculations
    quantity = random.choices([1, 2, 3, 4, 5], weights=[65, 20, 8, 5, 2])[0]
    total_amount = base_price * quantity

    # Location
    state = random.choices(list(townships_by_state.keys()), weights=[50, 20, 10, 8, 7, 5])[0]
    township = random.choice(townships_by_state[state])

    # Dates
    days_offset = random.randint(0, 100)
    order_date = start_date + timedelta(days=days_offset)

    # Shipping time 1-5 days
    ship_date = order_date + timedelta(days=random.randint(1, 5))

    payment = random.choice(payment_methods)
    channel = random.choice(channels)
    customer_name = f"Customer_{100 + i}"
    phone = f"09{random.randint(10000000, 99999999)}"

    status = random.choices(['Delivered', 'Cancelled', 'Returned'], weights=[88, 8, 4])[0]

    row = {
        'Order_ID': order_id,
        'Order_Date': order_date.strftime('%Y-%m-%d'),
        'Ship_Date': ship_date.strftime('%Y-%m-%d'),
        'Customer_Name': customer_name,
        'Phone_Number': phone,
        'Region_State': state,
        'Township': township,
        'Sales_Channel': channel,
        'Product_Category': category,
        'Product_Name': product_name,
        'Unit_Price_MMK': base_price,
        'Quantity': quantity,
        'Total_Amount_MMK': total_amount,
        'Payment_Method': payment,
        'Order_Status': status
    }
    data.append(row)

df = pd.DataFrame(data)

# --- INJECTING DIRTY DATA FOR CLEANING CHALLENGES ---

# 1. Missing Values (NaN / Nulls)
for _ in range(8):
    df.loc[random.randint(0, num_rows-1), 'Phone_Number'] = np.nan
for _ in range(6):
    df.loc[random.randint(0, num_rows-1), 'Ship_Date'] = np.nan
for _ in range(5):
    df.loc[random.randint(0, num_rows-1), 'Total_Amount_MMK'] = np.nan

# 2. Duplicate Rows (Exact duplicates and partial structural duplicates)
dup_indices = [12, 45, 78, 112]
df_dups = df.iloc[dup_indices].copy()
df = pd.concat([df, df_dups], ignore_index=True)

# 3. Inconsistent Text / Casing / Typos
for idx in random.sample(range(num_rows), 10):
    df.loc[idx, 'Product_Category'] = df.loc[idx, 'Product_Category'].lower()
for idx in random.sample(range(num_rows), 8):
    df.loc[idx, 'Payment_Method'] = df.loc[idx, 'Payment_Method'].replace('KBZPay', 'kbz pay').replace('WaveMoney', 'wave money')
for idx in random.sample(range(num_rows), 6):
    df.loc[idx, 'Region_State'] = df.loc[idx, 'Region_State'].replace('Yangon', 'Ygn').replace('Mandalay', 'MDY')

# 4. Outliers & Formatting Issues in Structural Columns
# Miscalculated total amount
df.loc[23, 'Total_Amount_MMK'] = 99999999
# Bad date formats
df.loc[14, 'Order_Date'] = '05/12/2026'
df.loc[55, 'Order_Date'] = '2026.03.15'
# Strange character spaces or trailing prefixes in phone numbers
df.loc[34, 'Phone_Number'] = '+959-12345678'
df.loc[88, 'Phone_Number'] = '09 444 555 66'

# Let's write this raw data to a neat multi-tab workbook with openpyxl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Sheet 1: ReadMe / Instructions
ws_readme = wb.active
ws_readme.title = "Data Cleaning Instructions"
ws_readme.views.sheetView[0].showGridLines = True

# Styling helpers
navy_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
light_grey = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
soft_accent = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_white_bold = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
font_normal = Font(name="Segoe UI", size=11)
font_bold = Font(name="Segoe UI", size=11, bold=True)
font_title = Font(name="Segoe UI", size=16, bold=True, color="1F4E78")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Build Readme Page
ws_readme.append([])
ws_readme.cell(row=2, column=2, value="Myanmar E-Commerce Sales Analysis").font = font_title
ws_readme.cell(row=3, column=2, value="Data Analyst Practical Assessment / Exercise File").font = font_normal
ws_readme.append([])

ws_readme.cell(row=5, column=2, value="Task Overview:").font = font_bold
ws_readme.cell(row=6, column=2, value="The 'Raw_Sales_Data' tab contains 150+ transactional records from a mock Myanmar e-commerce platform.").font = font_normal
ws_readme.cell(row=7, column=2, value="However, the data is 'dirty' and requires structural transformations and cleanup before running analysis or building dashboards.").font = font_normal

instructions = [
    "Identify and drop exact duplicate rows.",
    "Handle Missing Values: Find empty cells in 'Phone_Number', 'Ship_Date', and cross-calculate 'Total_Amount_MMK' where it is missing.",
    "Fix Casing & Text Inconsistency: Standardize 'Product_Category' and 'Payment_Method' variations (e.g. 'kbz pay' vs 'KBZPay', 'Ygn' vs 'Yangon').",
    "Standardize Date Formats: Ensure all entries in 'Order_Date' and 'Ship_Date' follow standard YYYY-MM-DD structures.",
    "Clean Phone Numbers: Standardize local mobile formats to standard '09xxxxxxxx' string format.",
    "Identify Outliers: Detect anomalous numeric entries like the impossible 'Total_Amount_MMK' value.",
    "Once cleaned, prepare a summary pivot table showing Total Sales Revenue by Region_State and Top Product Categories."
]

ws_readme.cell(row=9, column=2, value="Data Issues to Find & Fix:").font = font_bold
for idx, inst in enumerate(instructions, start=10):
    ws_readme.cell(row=idx, column=2, value=f"{idx-9}. {inst}").font = font_normal

# Sheet 2: Raw Data
ws_data = wb.create_sheet(title="Raw_Sales_Data")
ws_data.views.sheetView[0].showGridLines = True

# Write Header
headers = list(df.columns)
ws_data.append(headers)
for col_num, header in enumerate(headers, start=1):
    cell = ws_data.cell(row=1, column=col_num)
    cell.font = font_white_bold
    cell.fill = navy_header
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Write Rows
for r_idx, row in df.iterrows():
    row_values = list(row)
    # Convert numpy nan to None for Excel representation
    row_values = [None if pd.isna(v) else v for v in row_values]
    ws_data.append(row_values)

    # Optional styling for rows
    curr_row = ws_data.max_row
    for c_idx in range(1, len(headers) + 1):
        cell = ws_data.cell(row=curr_row, column=c_idx)
        cell.font = font_normal
        cell.border = thin_border

        # Zebra striping
        if curr_row % 2 == 0:
            cell.fill = light_grey

        # Alignments & Number formats
        if headers[c_idx-1] in ['Unit_Price_MMK', 'Total_Amount_MMK']:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
        elif headers[c_idx-1] in ['Quantity']:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="center")
        elif headers[c_idx-1] in ['Order_Date', 'Ship_Date', 'Order_ID', 'Phone_Number']:
            cell.alignment = Alignment(horizontal="center")

# Autofit Column Widths
for ws in [ws_readme, ws_data]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Specific manual adjustment for readme visibility
ws_readme.column_dimensions['B'].width = 110

# Save Workbook
filename = "Myanmar_Ecommerce_Sales_Dirty_Data8.xlsx"
wb.save(filename)
print(f"Workbook successfully saved to {filename}")
